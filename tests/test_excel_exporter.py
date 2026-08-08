"""
Unit tests for openpyxl Excel exporter functionality.
"""

import io
import openpyxl
import pandas as pd
import pytest
from src.excel_exporter import (
    dataframe_to_excel_bytes,
    export_multi_sheet_workbook,
    export_statements_zip
)


def test_dataframe_to_excel_bytes():
    data = {
        "Date": ["2024-01-15"],
        "Description": ["Tech Corp Payroll"],
        "Debit": [None],
        "Credit": [4250.00],
        "Amount": [4250.00],
        "Balance": [4250.00]
    }
    df = pd.DataFrame(data)
    excel_bytes = dataframe_to_excel_bytes(df, sheet_name="TestSheet")

    wb = openpyxl.load_workbook(io.BytesIO(excel_bytes))
    assert "TestSheet" in wb.sheetnames

    ws = wb["TestSheet"]
    assert ws.cell(row=1, column=1).value == "Date"
    assert ws.cell(row=1, column=4).value == "Credit"
    assert ws.cell(row=2, column=1).value == "2024-01-15"
    assert ws.cell(row=2, column=4).value == 4250.00
    assert ws.cell(row=2, column=4).number_format == "$#,##0.00;($#,##0.00);\"-\""


def test_export_multi_sheet_workbook():
    df1 = pd.DataFrame({"Date": ["2024-01-01"], "Amount": [100.00]})
    df2 = pd.DataFrame({"Date": ["2024-02-01"], "Amount": [200.00]})
    statements = {
        "Statement_Jan.pdf": df1,
        "Statement_Feb.pdf": df2
    }

    wb_bytes = export_multi_sheet_workbook(statements)
    wb = openpyxl.load_workbook(io.BytesIO(wb_bytes))

    assert "Statement_Jan" in wb.sheetnames
    assert "Statement_Feb" in wb.sheetnames

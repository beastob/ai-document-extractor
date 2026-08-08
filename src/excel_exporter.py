"""
Openpyxl Excel Exporter Module.
Formats extracted bank statement DataFrames into standardized Excel (.xlsx) files
with auto-fitted column widths, currency formatting ($#,##0.00), and date formatting (YYYY-MM-DD).
Supports batch packaging (ZIP, multi-sheet, or merged single sheet).
"""

import io
import zipfile
from typing import Dict, List, Optional, Tuple, Union
import pandas as pd

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Formatting string constants
CURRENCY_FORMAT = "$#,##0.00;($#,##0.00);\"-\""
DATE_FORMAT = "yyyy-mm-dd"


def write_excel_metadata_header(ws: openpyxl.worksheet.worksheet.Worksheet, metadata: Dict[str, str]) -> int:
    """
    Writes bank statement metadata summary rows at the top of the worksheet.
    Returns the row index where the main transaction table headers should start.
    """
    if not metadata:
        return 1

    header_title_font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
    meta_label_font = Font(name="Calibri", size=11, bold=True)
    meta_val_font = Font(name="Calibri", size=11, bold=False)
    navy_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    light_blue_fill = PatternFill(start_color="F0F4FF", end_color="F0F4FF", fill_type="solid")
    
    thin_border_side = Side(border_style="thin", color="D3D3D3")
    meta_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)

    # Row 1: Title Banner
    ws.merge_cells("A1:D1")
    title_cell = ws.cell(row=1, column=1, value="🏦 BANK STATEMENT ACCOUNT DETAILS")
    title_cell.font = header_title_font
    title_cell.fill = navy_fill
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 24

    # Row 2: Bank Name & Client Name
    a2 = ws.cell(row=2, column=1, value="Bank Name:")
    a2.font = meta_label_font
    a2.fill = light_blue_fill
    a2.border = meta_border

    b2 = ws.cell(row=2, column=2, value=metadata.get("Bank Name", "Unknown Bank"))
    b2.font = meta_val_font
    b2.border = meta_border

    c2 = ws.cell(row=2, column=3, value="Client Name:")
    c2.font = meta_label_font
    c2.fill = light_blue_fill
    c2.border = meta_border

    d2 = ws.cell(row=2, column=4, value=metadata.get("Client Name", "N/A"))
    d2.font = meta_val_font
    d2.border = meta_border

    # Row 3: BSB & Account Number
    a3 = ws.cell(row=3, column=1, value="BSB Number:")
    a3.font = meta_label_font
    a3.fill = light_blue_fill
    a3.border = meta_border

    b3 = ws.cell(row=3, column=2, value=metadata.get("BSB Number", "N/A"))
    b3.font = meta_val_font
    b3.border = meta_border

    c3 = ws.cell(row=3, column=3, value="Account Number:")
    c3.font = meta_label_font
    c3.fill = light_blue_fill
    c3.border = meta_border

    d3 = ws.cell(row=3, column=4, value=metadata.get("Account Number", "N/A"))
    d3.font = meta_val_font
    d3.border = meta_border

    return 5


def dataframe_to_excel_bytes(
    df: pd.DataFrame,
    sheet_name: str = "Bank Statement",
    metadata: Optional[Dict[str, str]] = None
) -> bytes:
    """
    Converts a standardized bank statement DataFrame into a single Excel file (.xlsx) in memory.
    Applies auto-fitted column widths, metadata summary block, and native openpyxl formatting.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]

    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    body_font = Font(name="Calibri", size=11, bold=False)
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")
    
    thin_border_side = Side(border_style="thin", color="D3D3D3")
    cell_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)

    # 1. Write metadata section if provided
    start_row = 1
    if metadata:
        start_row = write_excel_metadata_header(ws, metadata)

    # 2. Write transaction table headers
    headers = list(df.columns)
    for col_num, h in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col_num, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = cell_border
        if h in ["Debit", "Credit", "Amount", "Balance"]:
            cell.alignment = right_align
        elif h in ["Date", "Deductibility"]:
            cell.alignment = center_align
        else:
            cell.alignment = left_align

    # 3. Write data rows
    current_row = start_row + 1
    for r_idx, row in df.iterrows():
        for c_idx, col_name in enumerate(headers, 1):
            val = row[col_name]
            cell_val = None if (pd.isna(val) or val is None) else val
            cell = ws.cell(row=current_row, column=c_idx, value=cell_val)
            cell.font = body_font
            cell.border = cell_border

            if col_name in ["Debit", "Credit", "Amount", "Balance"]:
                cell.alignment = right_align
                if cell.value is not None and isinstance(cell.value, (int, float)):
                    cell.number_format = CURRENCY_FORMAT
            elif col_name in ["Date", "Deductibility"]:
                cell.alignment = center_align
                if col_name == "Date" and cell.value is not None:
                    cell.number_format = DATE_FORMAT
            else:
                cell.alignment = left_align
        current_row += 1

    # 4. Auto-fit column widths
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        hdr_cell = ws.cell(row=start_row, column=col[0].column)
        col_name = str(hdr_cell.value or "")

        for cell in col:
            val = cell.value
            if val is not None:
                if col_name in ["Debit", "Credit", "Amount", "Balance"] and isinstance(val, (int, float)):
                    val_str = f"${val:,.2f}"
                else:
                    val_str = str(val)
                max_len = max(max_len, len(val_str))

        adjusted_width = max(max_len + 4, 12)
        ws.column_dimensions[col_letter].width = adjusted_width

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def export_multi_sheet_workbook(
    statements: Dict[str, pd.DataFrame],
    metadata_dict: Optional[Dict[str, Dict[str, str]]] = None
) -> bytes:
    """
    Exports multiple statement DataFrames into a single Excel Workbook with multiple sheets.
    """
    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    body_font = Font(name="Calibri", size=11, bold=False)
    thin_border_side = Side(border_style="thin", color="D3D3D3")
    cell_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)

    for filename, df in statements.items():
        # Sanitize sheet title
        sheet_title = filename.replace(".pdf", "").replace(".PDF", "")
        sheet_title = "".join(c for c in sheet_title if c not in r"\/?*:[]")[:31]
        if not sheet_title:
            sheet_title = "Statement"

        existing = [s.title for s in wb.worksheets]
        orig_title = sheet_title
        counter = 1
        while sheet_title in existing:
            sheet_title = f"{orig_title[:27]}_{counter}"
            counter += 1

        ws = wb.create_sheet(title=sheet_title)

        meta = metadata_dict.get(filename) if metadata_dict else None
        start_row = 1
        if meta:
            start_row = write_excel_metadata_header(ws, meta)

        headers = list(df.columns)
        for col_num, h in enumerate(headers, 1):
            cell = ws.cell(row=start_row, column=col_num, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = cell_border
            if h in ["Debit", "Credit", "Amount", "Balance"]:
                cell.alignment = Alignment(horizontal="right", vertical="center")
            elif h in ["Date", "Deductibility"]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")

        current_row = start_row + 1
        for _, row in df.iterrows():
            for c_idx, col_name in enumerate(headers, 1):
                val = row[col_name]
                cell_val = None if (pd.isna(val) or val is None) else val
                cell = ws.cell(row=current_row, column=c_idx, value=cell_val)
                cell.font = body_font
                cell.border = cell_border
                if col_name in ["Debit", "Credit", "Amount", "Balance"]:
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = CURRENCY_FORMAT
                elif col_name in ["Date", "Deductibility"]:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    if col_name == "Date" and cell.value is not None:
                        cell.number_format = DATE_FORMAT
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center")
            current_row += 1

        # Auto-fit column widths
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            hdr_cell = ws.cell(row=start_row, column=col[0].column)
            col_name = str(hdr_cell.value or "")

            for cell in col:
                val = cell.value
                if val is not None:
                    if col_name in ["Debit", "Credit", "Amount", "Balance"] and isinstance(val, (int, float)):
                        val_str = f"${val:,.2f}"
                    else:
                        val_str = str(val)
                    max_len = max(max_len, len(val_str))
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def export_statements_zip(
    statements: Dict[str, pd.DataFrame],
    metadata_dict: Optional[Dict[str, Dict[str, str]]] = None
) -> bytes:
    """
    Exports a dictionary of statement DataFrames as a ZIP archive containing individual .xlsx files.
    """
    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zip_file:
        for filename, df in statements.items():
            excel_name = filename.rsplit(".", 1)[0] + "_extracted.xlsx"
            meta = metadata_dict.get(filename) if metadata_dict else None
            excel_bytes = dataframe_to_excel_bytes(df, sheet_name="Statement", metadata=meta)
            zip_file.writestr(excel_name, excel_bytes)

    return zip_buffer.getvalue()
